"""Universal file preview tool.

Preview any file content - text locally, images/PDF via Claude Vision API.

Supports multiple file sources:
- exec_* (temp_id): Files from execute_python in Redis cache
- file_* (claude_file_id): Files in Claude Files API
- Telegram file_id: Files from user uploads

For text formats (CSV, JSON, code) - FREE local parsing.
For images/PDF - PAID Claude Vision/PDF API call.

NO __init__.py - use direct import:
    from core.tools.preview_file import preview_file, PREVIEW_FILE_TOOL
"""

import asyncio
from typing import Any, Dict, Optional, TYPE_CHECKING

from anthropic import APIStatusError
from cache.exec_cache import get_exec_file
from cache.exec_cache import get_exec_meta
from core.clients import get_anthropic_client
from core.pricing import calculate_claude_cost
from core.pricing import cost_to_float
from db.repositories.user_file_repository import UserFileRepository
from utils.structured_logging import get_logger

if TYPE_CHECKING:
    from aiogram import Bot
    from sqlalchemy.ext.asyncio import AsyncSession

logger = get_logger(__name__)

# Retry configuration for Claude API
MAX_RETRIES = 3
RETRY_DELAY_SECONDS = 2.0
RETRYABLE_STATUS_CODES = {500, 502, 503, 504, 529}

# Model for vision/PDF analysis
VISION_MODEL_ID = "claude-sonnet-4-5-20250929"

PREVIEW_FILE_TOOL = {
    "name":
        "preview_file",
    "description":
        """YOUR internal verification tool — preview ANY file BEFORE delivering to user.

<purpose>
See what's in any file before sending to user. Does NOT deliver to user.
Use this to verify generated content matches user's request before calling deliver_file.
Works with ALL file types from ALL sources.
</purpose>

<file_sources>
- exec_xxx: Files from execute_python (images, PDFs, CSV, text, binary)
- file_xxx: Files already in Claude Files API
- Telegram file_id: Files from user messages

ALL sources supported — just pass the file_id.
</file_sources>

<processing_by_type>
- CSV/XLSX: Shows table with rows (FREE, local parsing)
- Text/JSON/XML/code: Shows content with line numbers (FREE)
- Images: Claude Vision describes what's visible (PAID)
  → For exec_xxx images, auto-uploads to Files API first
- PDF: Claude PDF analyzes document content (PAID)
  → For exec_xxx PDFs, auto-uploads to Files API first
- Audio/Video: Shows metadata, use transcribe_audio for content
</processing_by_type>

<verification_workflow>
1. execute_python/render_latex → output_files with temp_id
2. preview_file(file_id="exec_xxx", question="Does this show...?")
   → Your internal check, NOT sent to user
3. If correct: deliver_file(temp_id="exec_xxx") → sends to user
4. If wrong: regenerate, preview again, then deliver
</verification_workflow>

<examples>
1. Verify generated PDF before sending:
   preview_file(file_id="exec_abc123", question="Does this PDF have all 5 sections?")

2. Check chart data is correct:
   preview_file(file_id="exec_xyz789", question="Does the chart show Q1-Q4 2024 data?")

3. Verify CSV export:
   preview_file(file_id="exec_def456", max_rows=5)
</examples>

<cost>
FREE for text/CSV/XLSX (local parsing).
PAID for images/PDF (Claude Vision API, ~$0.003-0.01).
</cost>""",
    "input_schema": {
        "type": "object",
        "properties": {
            "file_id": {
                "type": "string",
                "description":
                    "File identifier: exec_xxx (from execute_python), "
                    "file_xxx (Files API), or Telegram file_id"
            },
            "question": {
                "type": "string",
                "description": "For images/PDF: what to look for. "
                               "Default: 'Describe the content of this file'"
            },
            "max_rows": {
                "type": "integer",
                "description": "Max rows for CSV/XLSX preview (default: 20)"
            },
            "max_chars": {
                "type": "integer",
                "description": "Max characters for text files (default: 5000)"
            }
        },
        "required": ["file_id"]
    }
}


async def _get_file_from_exec_cache(
    temp_id: str,) -> tuple[Optional[bytes], Optional[Dict[str, Any]]]:
    """Get file from execute_python Redis cache.

    Args:
        temp_id: Temporary file ID (exec_xxx).

    Returns:
        Tuple of (content bytes, metadata dict) or (None, None) if not found.
    """
    metadata = await get_exec_meta(temp_id)
    if not metadata:
        return None, None

    content = await get_exec_file(temp_id)
    if not content:
        return None, None

    return content, metadata


async def _get_file_from_db(
    file_id: str,
    session: 'AsyncSession',
    bot: 'Bot',
) -> tuple[Optional[bytes], Optional[Dict[str, Any]], Optional[str]]:
    """Get file from database (telegram or claude file_id).

    Args:
        file_id: Claude file_id or Telegram file_id.
        session: Database session.
        bot: Telegram bot for downloading.

    Returns:
        Tuple of (content bytes, metadata dict, claude_file_id) or
        (None, None, None) if not found.
    """
    user_file_repo = UserFileRepository(session)

    # Try to find by claude_file_id first
    user_file = await user_file_repo.get_by_claude_file_id(file_id)

    # If not found, try by telegram_file_id
    if not user_file:
        user_file = await user_file_repo.get_by_telegram_file_id(file_id)

    if not user_file:
        return None, None, None

    # Build metadata
    metadata = {
        "filename": user_file.filename,
        "mime_type": user_file.mime_type,
        "file_size": user_file.file_size,
        "file_type": user_file.file_type.value,
    }

    # Download content
    try:
        if user_file.telegram_file_id:
            # Download from Telegram
            tg_file = await bot.get_file(user_file.telegram_file_id)
            from io import BytesIO
            buffer = BytesIO()
            await bot.download_file(tg_file.file_path, buffer)
            content = buffer.getvalue()
        else:
            # Download from Files API
            from core.claude.files_api import download_from_files_api
            content = await download_from_files_api(user_file.claude_file_id)

        return content, metadata, user_file.claude_file_id

    except Exception as e:
        logger.warning("preview_file.download_failed",
                       file_id=file_id,
                       error=str(e))
        return None, None, None


def _is_text_format(mime_type: str, filename: str) -> bool:
    """Check if file is a text format that can be parsed locally."""
    text_mimes = {
        "text/plain",
        "text/csv",
        "text/html",
        "text/xml",
        "application/json",
        "application/xml",
        "application/javascript",
    }
    text_extensions = {
        ".txt",
        ".csv",
        ".json",
        ".xml",
        ".html",
        ".js",
        ".py",
        ".md",
        ".yaml",
        ".yml",
        ".toml",
        ".ini",
        ".cfg",
        ".log",
        ".sh",
        ".bash",
        ".sql",
        ".css",
        ".scss",
        ".less",
    }

    if mime_type in text_mimes or mime_type.startswith("text/"):
        return True

    return any(filename.lower().endswith(ext) for ext in text_extensions)


def _is_spreadsheet(mime_type: str, filename: str) -> bool:
    """Check if file is a spreadsheet (CSV/XLSX)."""
    spreadsheet_mimes = {
        "text/csv",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    }
    return (mime_type in spreadsheet_mimes or filename.lower().endswith(
        (".csv", ".xlsx", ".xls")))


def _is_image(mime_type: str) -> bool:
    """Check if file is an image."""
    return mime_type.startswith("image/")


def _is_pdf(mime_type: str, filename: str) -> bool:
    """Check if file is a PDF."""
    return mime_type == "application/pdf" or filename.lower().endswith(".pdf")


def _is_audio_video(mime_type: str) -> bool:
    """Check if file is audio or video."""
    return mime_type.startswith("audio/") or mime_type.startswith("video/")


def _preview_csv(
    content: bytes,
    metadata: Dict[str, Any],
    max_rows: int,
) -> Dict[str, Any]:
    """Preview CSV file content."""
    try:
        import csv
        import io

        try:
            text = content.decode("utf-8")
        except UnicodeDecodeError:
            text = content.decode("latin-1")

        reader = csv.reader(io.StringIO(text))
        rows = []
        for i, row in enumerate(reader):
            if i >= max_rows + 1:
                break
            rows.append(row)

        if not rows:
            return {"success": "false", "error": "Empty CSV file"}

        header = rows[0]
        data_rows = rows[1:]
        total_rows = text.count('\n')

        # Format as table
        col_widths = []
        for col_idx, col_name in enumerate(header):
            col_vals = [col_name] + [
                row[col_idx] if col_idx < len(row) else "" for row in data_rows
            ]
            max_width = min(30, max(len(str(v)) for v in col_vals))
            col_widths.append(max_width)

        def format_row(row):
            cells = []
            for i, cell in enumerate(row):
                width = col_widths[i] if i < len(col_widths) else 20
                cell_str = str(cell)[:width]
                cells.append(cell_str.ljust(width))
            return " | ".join(cells)

        table_lines = [format_row(header)]
        table_lines.append("-" * len(table_lines[0]))
        for row in data_rows:
            table_lines.append(format_row(row))

        return {
            "success": "true",
            "filename": metadata.get("filename"),
            "file_type": "csv",
            "columns": header,
            "column_count": len(header),
            "total_rows": total_rows,
            "previewed_rows": len(data_rows),
            "content": "\n".join(table_lines),
            "message": f"Showing {len(data_rows)} of ~{total_rows} rows"
        }

    except Exception as e:
        return {"success": "false", "error": f"CSV parse error: {str(e)}"}


def _preview_excel(
    content: bytes,
    metadata: Dict[str, Any],
    max_rows: int,
) -> Dict[str, Any]:
    """Preview Excel file content."""
    try:
        import io

        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        sheet = wb.active

        if sheet is None:
            return {"success": "false", "error": "No active sheet"}

        rows = []
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i >= max_rows + 1:
                break
            rows.append([str(cell) if cell is not None else "" for cell in row])

        wb.close()

        if not rows:
            return {"success": "false", "error": "Empty Excel file"}

        header = rows[0]
        data_rows = rows[1:]

        # Format as table
        col_widths = []
        for col_idx, col_name in enumerate(header):
            col_vals = [col_name] + [
                row[col_idx] if col_idx < len(row) else "" for row in data_rows
            ]
            max_width = min(30, max(len(str(v)) for v in col_vals))
            col_widths.append(max_width)

        def format_row(row):
            cells = []
            for i, cell in enumerate(row):
                width = col_widths[i] if i < len(col_widths) else 20
                cell_str = str(cell)[:width]
                cells.append(cell_str.ljust(width))
            return " | ".join(cells)

        table_lines = [format_row(header)]
        table_lines.append("-" * len(table_lines[0]))
        for row in data_rows:
            table_lines.append(format_row(row))

        return {
            "success": "true",
            "filename": metadata.get("filename"),
            "file_type": "xlsx",
            "sheet_name": sheet.title,
            "columns": header,
            "column_count": len(header),
            "total_rows": sheet.max_row or 0,
            "previewed_rows": len(data_rows),
            "content": "\n".join(table_lines),
            "message": f"Showing {len(data_rows)} of {sheet.max_row} rows"
        }

    except ImportError:
        return {"success": "false", "error": "openpyxl not installed"}
    except Exception as e:
        return {"success": "false", "error": f"Excel parse error: {str(e)}"}


def _preview_text(
    content: bytes,
    metadata: Dict[str, Any],
    max_chars: int,
) -> Dict[str, Any]:
    """Preview text file content."""
    try:
        try:
            text = content.decode("utf-8")
        except UnicodeDecodeError:
            text = content.decode("latin-1")

        total_chars = len(text)
        total_lines = text.count('\n') + 1
        truncated = total_chars > max_chars
        preview = text[:max_chars]

        # Add line numbers
        lines = preview.split('\n')
        numbered_lines = []
        for i, line in enumerate(lines, 1):
            if len(line) > 120:
                line = line[:117] + "..."
            numbered_lines.append(f"{i:4d} | {line}")

        return {
            "success":
                "true",
            "filename":
                metadata.get("filename"),
            "file_type":
                "text",
            "total_chars":
                total_chars,
            "total_lines":
                total_lines,
            "content":
                "\n".join(numbered_lines),
            "truncated":
                truncated,
            "message": (f"Showing {len(preview)} of {total_chars} chars"
                        if truncated else "Full content shown")
        }

    except Exception as e:
        return {"success": "false", "error": f"Text decode error: {str(e)}"}


async def _preview_image_with_vision(
    claude_file_id: str,
    question: str,
) -> Dict[str, Any]:
    """Preview image using Claude Vision API (PAID)."""
    client = get_anthropic_client(use_files_api=True)

    last_error = None
    for attempt in range(MAX_RETRIES):
        try:

            def _sync_call():
                return client.messages.create(
                    model=VISION_MODEL_ID,
                    max_tokens=4096,
                    messages=[{
                        "role":
                            "user",
                        "content": [{
                            "type": "image",
                            "source": {
                                "type": "file",
                                "file_id": claude_file_id
                            }
                        }, {
                            "type": "text",
                            "text": question
                        }]
                    }])

            response = await asyncio.to_thread(_sync_call)

            analysis = response.content[0].text
            input_tokens = response.usage.input_tokens
            output_tokens = response.usage.output_tokens

            cost_usd = calculate_claude_cost(
                model_id=VISION_MODEL_ID,
                input_tokens=input_tokens,
                output_tokens=output_tokens,
            )

            logger.info("preview_file.vision_success",
                        claude_file_id=claude_file_id,
                        tokens=input_tokens + output_tokens,
                        cost_usd=cost_to_float(cost_usd))

            return {
                "success": "true",
                "file_type": "image",
                "content": analysis,
                "tokens_used": str(input_tokens + output_tokens),
                "cost_usd": f"{cost_to_float(cost_usd):.6f}",
                "message": "Image analyzed with Claude Vision",
                # Detailed token info for cost tracking
                "_model_id": VISION_MODEL_ID,
                "_input_tokens": input_tokens,
                "_output_tokens": output_tokens,
                "_cache_read_tokens": 0,
                "_cache_creation_tokens": 0,
            }

        except APIStatusError as e:
            last_error = e
            if e.status_code in RETRYABLE_STATUS_CODES:
                if attempt < MAX_RETRIES - 1:
                    await asyncio.sleep(RETRY_DELAY_SECONDS * (2**attempt))
                    continue
            raise

    if last_error:
        raise last_error
    raise RuntimeError("Unexpected: retry loop completed without result")


async def _preview_pdf_with_vision(
    claude_file_id: str,
    question: str,
) -> Dict[str, Any]:
    """Preview PDF using Claude PDF API (PAID)."""
    client = get_anthropic_client(use_files_api=True)

    last_error = None
    for attempt in range(MAX_RETRIES):
        try:

            def _sync_call():
                return client.messages.create(
                    model=VISION_MODEL_ID,
                    max_tokens=4096,
                    messages=[{
                        "role":
                            "user",
                        "content": [{
                            "type": "document",
                            "source": {
                                "type": "file",
                                "file_id": claude_file_id
                            }
                        }, {
                            "type": "text",
                            "text": question
                        }]
                    }])

            response = await asyncio.to_thread(_sync_call)

            analysis = response.content[0].text
            input_tokens = response.usage.input_tokens
            output_tokens = response.usage.output_tokens

            cost_usd = calculate_claude_cost(
                model_id=VISION_MODEL_ID,
                input_tokens=input_tokens,
                output_tokens=output_tokens,
            )

            logger.info("preview_file.pdf_success",
                        claude_file_id=claude_file_id,
                        tokens=input_tokens + output_tokens,
                        cost_usd=cost_to_float(cost_usd))

            return {
                "success": "true",
                "file_type": "pdf",
                "content": analysis,
                "tokens_used": str(input_tokens + output_tokens),
                "cost_usd": f"{cost_to_float(cost_usd):.6f}",
                "message": "PDF analyzed with Claude",
                # Detailed token info for cost tracking
                "_model_id": VISION_MODEL_ID,
                "_input_tokens": input_tokens,
                "_output_tokens": output_tokens,
                "_cache_read_tokens": 0,
                "_cache_creation_tokens": 0,
            }

        except APIStatusError as e:
            last_error = e
            if e.status_code in RETRYABLE_STATUS_CODES:
                if attempt < MAX_RETRIES - 1:
                    await asyncio.sleep(RETRY_DELAY_SECONDS * (2**attempt))
                    continue
            raise

    if last_error:
        raise last_error
    raise RuntimeError("Unexpected: retry loop completed without result")


async def _preview_by_type(
    content: bytes,
    metadata: Dict[str, Any],
    mime_type: str,
    filename: str,
    claude_file_id: Optional[str],
    file_id: str,
    question: str,
    max_rows: int,
    max_chars: int,
) -> Dict[str, Any]:
    """Route preview to appropriate handler based on file type.

    Args:
        content: File content bytes.
        metadata: File metadata dict.
        mime_type: MIME type of file.
        filename: Original filename.
        claude_file_id: Claude Files API ID if available.
        file_id: Original file_id for logging.
        question: Question for image/PDF analysis.
        max_rows: Maximum rows for tabular data.
        max_chars: Maximum characters for text files.

    Returns:
        Dict with preview result.
    """
    # pylint: disable=too-many-return-statements
    # 1. Spreadsheets (CSV/XLSX) - FREE
    if _is_spreadsheet(mime_type, filename):
        if mime_type == "text/csv" or filename.endswith(".csv"):
            result = _preview_csv(content, metadata, max_rows)
        else:
            result = _preview_excel(content, metadata, max_rows)
        result["filename"] = filename
        return result

    # 2. Text files - FREE
    if _is_text_format(mime_type, filename):
        result = _preview_text(content, metadata, max_chars)
        result["filename"] = filename
        return result

    # 3. Images - PAID (Claude Vision)
    if _is_image(mime_type):
        return await _handle_image_preview(content, filename, mime_type,
                                           claude_file_id, file_id, question)

    # 4. PDF - PAID (Claude PDF)
    if _is_pdf(mime_type, filename):
        return await _handle_pdf_preview(content, filename, mime_type,
                                         claude_file_id, file_id, question)

    # 5. Audio/Video - FREE (metadata only)
    if _is_audio_video(mime_type):
        return {
            "success": "true",
            "filename": filename,
            "file_type": "audio_video",
            "mime_type": mime_type,
            "size_bytes": len(content),
            "content": f"Audio/video file: {filename} ({mime_type}, "
                       f"{len(content)} bytes). Use transcribe_audio tool "
                       "to get the spoken content.",
            "message": "Use transcribe_audio for audio/video content"
        }

    # 6. Unknown binary - FREE (info only, suggest execute_python)
    return _preview_binary(content, filename, mime_type)


async def _handle_image_preview(
    content: bytes,
    filename: str,
    mime_type: str,
    claude_file_id: Optional[str],
    file_id: str,
    question: str,
) -> Dict[str, Any]:
    """Handle image preview via Claude Vision API."""
    if not claude_file_id:
        from core.claude.files_api import upload_to_files_api
        claude_file_id = await upload_to_files_api(content, filename, mime_type)
        logger.info("preview_file.uploaded_for_vision",
                    filename=filename,
                    claude_file_id=claude_file_id)

    try:
        result = await _preview_image_with_vision(claude_file_id, question)
        result["filename"] = filename
        return result
    except Exception as e:
        logger.error("preview_file.vision_failed",
                     file_id=file_id,
                     error=str(e))
        return {
            "success": "false",
            "filename": filename,
            "error": f"Vision API error: {str(e)}"
        }


async def _handle_pdf_preview(
    content: bytes,
    filename: str,
    mime_type: str,
    claude_file_id: Optional[str],
    file_id: str,
    question: str,
) -> Dict[str, Any]:
    """Handle PDF preview via Claude PDF API."""
    if not claude_file_id:
        from core.claude.files_api import upload_to_files_api
        claude_file_id = await upload_to_files_api(content, filename, mime_type)
        logger.info("preview_file.uploaded_for_pdf",
                    filename=filename,
                    claude_file_id=claude_file_id)

    try:
        result = await _preview_pdf_with_vision(claude_file_id, question)
        result["filename"] = filename
        return result
    except Exception as e:
        logger.error("preview_file.pdf_failed", file_id=file_id, error=str(e))
        return {
            "success": "false",
            "filename": filename,
            "error": f"PDF API error: {str(e)}"
        }


def _preview_binary(
    content: bytes,
    filename: str,
    mime_type: str,
) -> Dict[str, Any]:
    """Preview unknown binary file with library suggestions."""
    ext = filename.lower().split('.')[-1] if '.' in filename else ""
    lib_suggestions = {
        "docx": "python-docx",
        "doc": "python-docx or antiword",
        "pptx": "python-pptx",
        "xlsx": "openpyxl or pandas",
        "xls": "xlrd or pandas",
        "sqlite": "sqlite3",
        "db": "sqlite3",
        "parquet": "pandas or pyarrow",
        "pkl": "pickle",
        "pickle": "pickle",
        "zip": "zipfile",
        "tar": "tarfile",
        "gz": "gzip",
        "7z": "py7zr",
        "rar": "rarfile",
        "npy": "numpy",
        "npz": "numpy",
        "h5": "h5py",
        "hdf5": "h5py",
        "mat": "scipy.io",
        "feather": "pandas or pyarrow",
        "avro": "fastavro",
        "msgpack": "msgpack",
        "bson": "bson",
    }

    suggestion = lib_suggestions.get(ext, "")
    if suggestion:
        parse_hint = (f"To parse this file, use execute_python with the "
                      f"'{suggestion}' library. The file is available at the "
                      f"path provided in input_files.")
    else:
        parse_hint = ("To inspect this file, use execute_python. Try reading "
                      "first bytes to detect format, or use 'file' command.")

    return {
        "success": "true",
        "filename": filename,
        "file_type": "binary",
        "mime_type": mime_type,
        "file_extension": ext,
        "size_bytes": len(content),
        "content":
            f"Binary file: {filename} ({mime_type}, {len(content)} bytes).\n\n"
            f"{parse_hint}",
        "message": "Use execute_python to parse binary files",
        "suggested_library": suggestion if suggestion else None,
    }


async def preview_file(
    file_id: str,
    bot: 'Bot',
    session: 'AsyncSession',
    thread_id: int | None = None,  # pylint: disable=unused-argument
    question: str = "Describe the content of this file",
    max_rows: int = 20,
    max_chars: int = 5000,
) -> Dict[str, Any]:
    """Preview file content from any source.

    Args:
        file_id: File identifier (exec_xxx, file_xxx, or Telegram file_id).
        bot: Telegram bot instance.
        session: Database session.
        thread_id: Thread ID (unused, for interface consistency).
        question: Question for image/PDF analysis.
        max_rows: Maximum rows for tabular data.
        max_chars: Maximum characters for text files.

    Returns:
        Dict with file content preview, optionally with cost_usd for paid ops.
    """
    logger.info("tools.preview_file.called",
                file_id=file_id,
                question_length=len(question),
                max_rows=max_rows,
                max_chars=max_chars)

    content: Optional[bytes] = None
    metadata: Optional[Dict[str, Any]] = None
    claude_file_id: Optional[str] = None

    # Determine source and get file
    if file_id.startswith("exec_"):
        content, metadata = await _get_file_from_exec_cache(file_id)
        if not content:
            return {
                "success": "false",
                "error": f"File '{file_id}' not found or expired (30 min TTL)"
            }
    else:
        content, metadata, claude_file_id = await _get_file_from_db(
            file_id, session, bot)
        if not content:
            return {
                "success": "false",
                "error": f"File '{file_id}' not found in database"
            }

    mime_type = metadata.get("mime_type", "application/octet-stream")
    filename = metadata.get("filename", "unknown")

    logger.debug("preview_file.processing",
                 file_id=file_id,
                 filename=filename,
                 mime_type=mime_type,
                 size_bytes=len(content),
                 has_claude_id=bool(claude_file_id))

    return await _preview_by_type(content, metadata, mime_type, filename,
                                  claude_file_id, file_id, question, max_rows,
                                  max_chars)


def _format_success_result(result: Dict[str, Any]) -> str:
    """Format successful preview result."""
    filename = result.get("filename", "file")
    file_type = result.get("file_type", "")
    cost = result.get("cost_usd")

    if file_type in ("csv", "xlsx"):
        rows = result.get("previewed_rows", 0)
        cols = result.get("column_count", 0)
        return f"[📋 {filename}: {rows} rows × {cols} cols]"

    if file_type == "text":
        lines = result.get("total_lines", 0)
        return f"[📄 {filename}: {lines} lines]"

    if file_type == "image":
        cost_str = f" ${cost}" if cost else ""
        return f"[🖼️ {filename}: analyzed{cost_str}]"

    if file_type == "pdf":
        cost_str = f" ${cost}" if cost else ""
        return f"[📑 {filename}: analyzed{cost_str}]"

    if file_type == "audio_video":
        return f"[🎵 {filename}: use transcribe_audio]"

    return f"[📁 {filename}]"


def format_preview_file_result(
    tool_input: Dict[str, Any],
    result: Dict[str, Any],
) -> str:
    """Format preview_file result for user display."""
    _ = tool_input

    if result.get("success") == "true":
        return _format_success_result(result)

    error = result.get("error", "unknown error")[:50]
    return f"[❌ Preview failed: {error}]"


# Unified tool configuration
from core.tools.base import ToolConfig

TOOL_CONFIG = ToolConfig(
    name="preview_file",
    definition=PREVIEW_FILE_TOOL,
    executor=preview_file,
    emoji="👁️",
    needs_bot_session=True,
    format_result=format_preview_file_result,
)
